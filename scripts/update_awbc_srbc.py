#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Update AWBC/SRBC prompts and fuse summaries/interpretations in multi-sheet Excel files.
"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
from typing import List, Optional, Sequence, Tuple

from openpyxl import load_workbook


AWBC_CONFIG_V1 = {
    "prompt": (
        "AWBC↑提示存在异常白细胞群或异常细胞成分增多，"
        "需结合血涂片及临床情况评估感染、炎症或血液系统异常。"
    ),
    "basis": (
        "AWBC为机器学习识别的异常白细胞计数，升高提示异常细胞聚集或未分类白细胞成分增加，"
        "建议复查或人工镜检。"
    ),
    "short": "AWBC↑提示异常白细胞风险",
    "interpretation": (
        "AWBC为异常白细胞计数，升高提示异常细胞成分增多，"
        "需结合感染/炎症或血液系统异常进一步评估。"
    ),
    "disease": {
        "name": "血液系统异常风险（需排除白血病/骨髓增生性疾病）",
        "probability": "中等",
        "analysis": (
            "AWBC↑提示存在异常白细胞群或未分类细胞成分增多，"
            "需结合血涂片或进一步检查排除血液系统异常。"
        ),
        "priority": 2,
        "keywords": ["AWBC", "异常白细胞"],
    },
}

SRBC_CONFIG_V1 = {
    "prompt": "SRBC↑提示镰刀型红细胞存在，需考虑镰刀型细胞贫血风险或携带者状态。",
    "basis": (
        "SRBC为镰刀型红细胞识别计数，升高提示红细胞形态异常，"
        "常见于镰刀型细胞贫血或携带者。"
    ),
    "short": "SRBC↑提示镰刀型细胞贫血风险",
    "interpretation": (
        "SRBC升高提示镰刀型红细胞存在，需结合溶血指标与病史评估镰刀型细胞贫血或携带者状态。"
    ),
    "disease": {
        "name": "镰刀型细胞贫血/携带者状态",
        "probability": "较高",
        "analysis": (
            "SRBC↑提示红细胞形态异常，需排查镰刀型细胞贫血或携带者状态，"
            "结合HGB/RET%等指标评估溶血程度。"
        ),
        "priority": 1,
        "keywords": ["SRBC", "镰刀"],
    },
}

REFERENCES_V2 = {
    "awbc": [
        "[9] Palmer L, et al. ICSH recommendations for the standardization of nomenclature and grading of peripheral blood cell morphological features. Int J Lab Hematol. 2015.",
        "[10] ISLH. Consensus rules for blood smear review / positive smear findings definitions.",
        "[11] CLSI. H20: Reference Leukocyte (WBC) Differential Count and Evaluation of Instrumental Methods.",
        "[12] Tripathi AK, et al. Laboratory Evaluation of Acute Leukemia. StatPearls. 2025.",
        "[13] Lynch EC. Peripheral Blood Smear. Clinical Methods (NCBI Bookshelf). 1990.",
        "[14] Al-Gwaiz LA, et al. The diagnostic value of toxic granulation in predicting bacterial infections. 2007."
    ],
    "srbc": [
        "[15] NHLBI (NIH). Sickle Cell Disease – Diagnosis. 2024.",
        "[16] NHLBI. Evidence-Based Management of Sickle Cell Disease (Guideline/Report).",
        "[17] American Society of Hematology. Hemoglobin Electrophoresis in Sickle Cell Disease.",
        "[18] Palmer L, et al. ICSH morphology recommendations. 2015.",
        "[19] National Academies. Protocols for screening should use hemoglobin electrophoresis or reliable separation methods."
    ]
}

AWBC_CONFIG_V2 = {
    "prompt": (
        "检测到白细胞AWBC指标升高异常，多见于：炎症/感染相关的毒性改变或异常形态细胞 [13][14]，"
        "外周血异常/幼稚细胞增多或血液系统异常线索 [9][12]，需进行涂片复核与异常形态确认 [10][11]。"
        "请医生查看细胞图片，建议进一步询问感染症状/近期用药史，或做外周血涂片复核与人工分类排查 [10][11]。"
    ),
    "basis": (
        "AWBC为异常白细胞计数，升高提示异常细胞聚集或未分类白细胞成分增加，"
        "建议复查或人工镜检。"
    ),
    "short": "AWBC↑提示异常白细胞风险",
    "interpretation": (
        "AWBC为异常白细胞计数，升高提示异常细胞成分增多，需结合感染/炎症或血液系统异常进一步评估。"
    ),
    "disease": {
        "name": "血液系统异常风险（需排除白血病/骨髓增生性疾病）",
        "probability": "中等",
        "analysis": (
            "AWBC↑提示存在异常白细胞群或未分类细胞成分增多，"
            "需结合血涂片或进一步检查排除血液系统异常。"
        ),
        "priority": 2,
        "keywords": ["AWBC", "异常白细胞"],
    },
}

SRBC_CONFIG_V2 = {
    "prompt": (
        "检测到红细胞SRBC指标升高异常，多见于：镰刀型细胞贫血或携带者状态 [15][16]，"
        "或镰刀红细胞形态提示 [18]。请医生查看细胞图片，建议进一步询问家族史/贫血相关症状，"
        "或做血红蛋白电泳/HPLC/基因检测 [15][17][19] 排查。"
    ),
    "basis": (
        "SRBC为镰刀型红细胞识别计数，升高提示红细胞形态异常，"
        "常见于镰刀型细胞贫血或携带者。"
    ),
    "short": "SRBC↑提示镰刀型细胞贫血风险",
    "interpretation": (
        "SRBC升高提示镰刀型红细胞存在，需结合溶血指标与病史评估镰刀型细胞贫血或携带者状态。"
    ),
    "disease": {
        "name": "镰刀型细胞贫血/携带者状态",
        "probability": "较高",
        "analysis": (
            "SRBC↑提示红细胞形态异常，需排查镰刀型细胞贫血或携带者状态，"
            "结合HGB/RET%等指标评估溶血程度。"
        ),
        "priority": 1,
        "keywords": ["SRBC", "镰刀"],
    },
}

CONFIG_PROFILES = {
    "v1": (AWBC_CONFIG_V1, SRBC_CONFIG_V1),
    "v2": (AWBC_CONFIG_V2, SRBC_CONFIG_V2),
}

NEUTRAL_PHRASES = ["无明显异常", "正常范围", "未见明显异常"]

ENGLISH_REPLACEMENTS: List[Tuple[re.Pattern, str]] = [
    (re.compile(r"\bHematology System abnormalities\b", re.IGNORECASE), "hematological abnormalities"),
    (re.compile(r"\banti[- ]?metabolic\b", re.IGNORECASE), "antimetabolite"),
    (re.compile(r"\batypical granulocytes\s*\(ALY#\)", re.IGNORECASE), "atypical lymphocytes (ALY#)"),
    (re.compile(r"\bimmature granulocytes\s*\(ALY#\)", re.IGNORECASE), "atypical lymphocytes (ALY#)"),
    (re.compile(r"\bgranulocytes\s*\(ALY#\)", re.IGNORECASE), "atypical lymphocytes (ALY#)"),
]

ALY_ADJECTIVES = r"(?:abnormal|atypical|unclassified|immature|degenerated|reactive|early)"
ALY_TARGET = r"(?:granulocytes?|white\s+blood\s+cells?|wbc(?:s)?|monocytes?|nucleated\s+cells?|cells?|components?)"
ALY_FIXES: List[Tuple[re.Pattern, str]] = [
    (
        re.compile(
            rf"\b(ALY#?)\s*\(\s*{ALY_ADJECTIVES}(?:\s+{ALY_ADJECTIVES})*\s+{ALY_TARGET}\s*\)",
            re.IGNORECASE,
        ),
        r"\1 (atypical lymphocytes)",
    ),
    (
        re.compile(
            rf"\b(ALY#?)\s*\(\s*{ALY_TARGET}\s*\)",
            re.IGNORECASE,
        ),
        r"\1 (atypical lymphocytes)",
    ),
    (
        re.compile(
            rf"\b{ALY_ADJECTIVES}(?:\s+{ALY_ADJECTIVES})*\s+{ALY_TARGET}\s*\(\s*(ALY#?)\s*\)",
            re.IGNORECASE,
        ),
        r"atypical lymphocytes (\1)",
    ),
    (
        re.compile(
            rf"\b{ALY_TARGET}\s*\(\s*(ALY#?)\s*\)",
            re.IGNORECASE,
        ),
        r"atypical lymphocytes (\1)",
    ),
]


def contains_any(text: str, keywords: Sequence[str]) -> bool:
    if not text:
        return False
    return any(keyword in text for keyword in keywords)


def cleanup(text: str) -> str:
    if not text:
        return text
    return text.replace("。；", "；").replace("；；", "；")


def cleanup_english(text: str) -> str:
    if not text:
        return text
    value = text
    for pattern, replacement in ENGLISH_REPLACEMENTS:
        value = pattern.sub(replacement, value)
    value = re.sub(r"\be\s*\.\s*g\s*\.\s*,", "e.g.,", value, flags=re.IGNORECASE)
    value = re.sub(r"\be\s*\.\s*g\s*\.\b", "e.g.", value, flags=re.IGNORECASE)
    return value


def cleanup_english_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cleaned = cleanup_english(cell.value)
                if cleaned != cell.value:
                    cell.value = cleaned


def fix_aly_descriptions(text: str) -> str:
    if not text or "ALY" not in text.upper():
        return text
    value = text
    for pattern, replacement in ALY_FIXES:
        value = pattern.sub(replacement, value)
    return value


def fix_aly_sheet(ws) -> None:
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str):
                cleaned = fix_aly_descriptions(cell.value)
                if cleaned != cell.value:
                    cell.value = cleaned


def prefix_summary(text: str, prefix: str, keywords: Sequence[str]) -> str:
    raw = text or ""
    if contains_any(raw, keywords):
        return raw
    if not raw:
        return prefix
    return cleanup(f"{prefix}；{raw}")


def prefix_short(text: str, prefix: str, keywords: Sequence[str]) -> str:
    raw = text or ""
    if contains_any(raw, keywords):
        return raw
    if not raw:
        return prefix
    return cleanup(f"{prefix}；{raw}")


def prefix_interpretation(text: str, prefix: str, keywords: Sequence[str]) -> str:
    raw = text or ""
    if contains_any(raw, keywords):
        return raw
    if not raw:
        return prefix
    return cleanup(f"{prefix} {raw}")


def insert_disease(
    row,
    start_col: int,
    name: str,
    probability: str,
    analysis: str,
    priority: int,
    keywords: Sequence[str],
) -> None:
    def get_slot(offset: int):
        base = start_col + offset * 3
        return row[base], row[base + 1], row[base + 2]

    def is_empty(cell):
        return cell.value is None or (isinstance(cell.value, str) and not cell.value.strip())

    # Already present
    slot_values = [get_slot(0)[0].value, get_slot(1)[0].value, get_slot(2)[0].value]
    if any(isinstance(val, str) and contains_any(val, keywords) for val in slot_values):
        return

    if priority == 1:
        # Shift down to make room at slot1
        slot3 = get_slot(2)
        slot2 = get_slot(1)
        slot1 = get_slot(0)

        dropped = (slot3[0].value, slot3[1].value, slot3[2].value)

        slot3[0].value, slot3[1].value, slot3[2].value = (
            slot2[0].value,
            slot2[1].value,
            slot2[2].value,
        )
        slot2[0].value, slot2[1].value, slot2[2].value = (
            slot1[0].value,
            slot1[1].value,
            slot1[2].value,
        )
        slot1[0].value = name
        slot1[1].value = probability
        slot1[2].value = analysis

        # Preserve dropped info by appending to slot3 analysis
        if any(dropped):
            tail = slot3[2].value or ""
            extra = dropped[2] or ""
            if extra:
                slot3[2].value = cleanup(f"{tail}；{extra}") if tail else extra
        return

    # Priority 2: fill first empty slot, else append analysis
    for offset in range(3):
        name_cell, prob_cell, analysis_cell = get_slot(offset)
        if is_empty(name_cell):
            name_cell.value = name
            prob_cell.value = probability
            analysis_cell.value = analysis
            return

    analysis_cell = get_slot(2)[2]
    analysis_cell.value = cleanup(f"{analysis_cell.value}；{analysis}") if analysis_cell.value else analysis


def update_sheet(ws, awbc_config, srbc_config) -> None:
    headers = [cell.value for cell in ws[1]]

    awbc_col = headers.index("AWBC#") + 1 if "AWBC#" in headers else None
    srbc_col = headers.index("SRBC#") + 1 if "SRBC#" in headers else None

    summary1_col = headers.index("总结1") + 1 if "总结1" in headers else None
    summary2_col = headers.index("总结2") + 1 if "总结2" in headers else None
    interp_col = headers.index("解读") + 1 if "解读" in headers else None
    disease1_col = headers.index("可能疾病1") + 1 if "可能疾病1" in headers else None

    for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
        if awbc_col:
            status = row[awbc_col - 1].value
            if status == "↑":
                prompt_cell = row[awbc_col] if awbc_col < len(row) else None
                basis_cell = row[awbc_col + 1] if awbc_col + 1 < len(row) else None
                if prompt_cell and not str(prompt_cell.value or "").strip():
                    prompt_cell.value = awbc_config["prompt"]
                if basis_cell and not str(basis_cell.value or "").strip():
                    basis_cell.value = awbc_config["basis"]

                keywords = awbc_config["disease"]["keywords"]
                if summary1_col:
                    cell = row[summary1_col - 1]
                    cell.value = prefix_summary(str(cell.value or ""), awbc_config["prompt"], keywords)
                if summary2_col:
                    cell = row[summary2_col - 1]
                    cell.value = prefix_short(str(cell.value or ""), awbc_config["short"], keywords)
                if interp_col:
                    cell = row[interp_col - 1]
                    cell.value = prefix_interpretation(
                        str(cell.value or ""), awbc_config["interpretation"], keywords
                    )
                if disease1_col:
                    insert_disease(
                        row,
                        disease1_col - 1,
                        awbc_config["disease"]["name"],
                        awbc_config["disease"]["probability"],
                        awbc_config["disease"]["analysis"],
                        awbc_config["disease"]["priority"],
                        keywords,
                    )

        if srbc_col:
            status = row[srbc_col - 1].value
            if status == "↑":
                prompt_cell = row[srbc_col] if srbc_col < len(row) else None
                basis_cell = row[srbc_col + 1] if srbc_col + 1 < len(row) else None
                if prompt_cell and not str(prompt_cell.value or "").strip():
                    prompt_cell.value = srbc_config["prompt"]
                if basis_cell and not str(basis_cell.value or "").strip():
                    basis_cell.value = srbc_config["basis"]

                keywords = srbc_config["disease"]["keywords"]
                if summary1_col:
                    cell = row[summary1_col - 1]
                    cell.value = prefix_summary(str(cell.value or ""), srbc_config["prompt"], keywords)
                if summary2_col:
                    cell = row[summary2_col - 1]
                    cell.value = prefix_short(str(cell.value or ""), srbc_config["short"], keywords)
                if interp_col:
                    cell = row[interp_col - 1]
                    cell.value = prefix_interpretation(
                        str(cell.value or ""), srbc_config["interpretation"], keywords
                    )
                if disease1_col:
                    insert_disease(
                        row,
                        disease1_col - 1,
                        srbc_config["disease"]["name"],
                        srbc_config["disease"]["probability"],
                        srbc_config["disease"]["analysis"],
                        srbc_config["disease"]["priority"],
                        keywords,
                    )


def resolve_output_path(input_path: Path, output_path: Optional[Path]) -> Path:
    if output_path:
        return output_path
    return input_path.with_name(
        f"{input_path.stem}_补充提示{input_path.suffix}"
    )


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Update AWBC/SRBC prompts and interpretations in Excel sheets."
    )
    parser.add_argument("--input", required=True, help="Input Excel file (.xlsx)")
    parser.add_argument("--output", help="Output Excel file (.xlsx)")
    parser.add_argument(
        "--profile",
        choices=CONFIG_PROFILES.keys(),
        default="v2",
        help="Prompt template profile (v1 keeps original wording, v2 includes citations).",
    )
    parser.add_argument(
        "--cleanup-english",
        action="store_true",
        help="Clean English phrases (ALY#, hematology phrasing, antimetabolite, e.g.).",
    )
    parser.add_argument(
        "--fix-aly",
        action="store_true",
        help="Fix ALY descriptions when paired with granulocytes or generic cell labels.",
    )
    parser.add_argument(
        "--skip-awbc-srbc",
        action="store_true",
        help="Skip AWBC/SRBC prompt updates; useful for cleanup-only runs.",
    )
    args = parser.parse_args()

    input_path = Path(args.input).expanduser()
    if not input_path.exists():
        raise SystemExit(f"Input file not found: {input_path}")
    output_path = resolve_output_path(
        input_path, Path(args.output).expanduser() if args.output else None
    )

    wb = load_workbook(input_path)
    for ws in wb.worksheets:
        if not args.skip_awbc_srbc:
            awbc_config, srbc_config = CONFIG_PROFILES[args.profile]
            update_sheet(ws, awbc_config, srbc_config)
        if args.cleanup_english:
            cleanup_english_sheet(ws)
        if args.fix_aly:
            fix_aly_sheet(ws)

    wb.save(output_path)
    print(f"Saved: {output_path}")


if __name__ == "__main__":
    main()
